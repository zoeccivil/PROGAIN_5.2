"""
Professional Report Generator for PROGRAIN 5.0

Generates PDF and Excel reports with unified modern styling and attachment support. 

Dependencies:
- Required: pandas, openpyxl, fpdf, PIL
- Optional: pdf2image (for embedding PDFs as images in reports)
"""

# type: ignore - Para pdf2image (dependencia opcional)

import pandas as pd
import io
import os
import logging
from fpdf import FPDF
from openpyxl. styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfWriter, PdfReader
import io
from PIL import Image
from pypdf import PdfWriter, PdfReader

logger = logging.getLogger(__name__)


class PDF(FPDF):
    def __init__(self, orientation='L', unit='mm', format='Letter', title="", project_name="", date_range=""):
        super().__init__(orientation, unit, format)
        self.report_title = title
        self. project_name = project_name
        self.date_range = date_range

    def header(self):
        # Encabezado moderno unificado para todos los reportes
        try:
            self.set_font('Arial', 'B', 16)
        except:
            self.set_font('Helvetica', 'B', 16)
            
        # Color título principal (Azul oscuro profesional)
        self.set_text_color(44, 62, 80)
        self.cell(0, 10, self.report_title, 0, 1, 'C')
        
        try:
            self.set_font('Arial', 'I', 10)
        except:
            self.set_font('Helvetica', 'I', 10)
            
        # Subtítulos
        self.set_text_color(100, 100, 100)  # Gris
        self.cell(0, 6, f"Proyecto: {self.project_name}", 0, 1, 'C')
        self.cell(0, 6, f"Período: {self.date_range}", 0, 1, 'C')
        self.ln(8)
        
        # Línea separadora
        self.set_draw_color(200, 200, 200)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        try:
            self.set_font('Arial', 'I', 8)
        except:
            self.set_font('Helvetica', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')


class ReportGenerator:
    def __init__(self, data=None, title="", project_name="", date_range="", currency_symbol="RD$", 
                 column_map=None, firebase_client=None, proyecto_id=None):
        """
        Initialize report generator with optional attachment support.
        
        Args:
            data: Transaction data
            title: Report title
            project_name: Project name
            date_range: Date range string
            currency_symbol: Currency symbol (default "RD$")
            column_map: Column mapping dictionary
            firebase_client: Firebase client for downloading attachments
            proyecto_id:  Project ID for attachments
        """
        self. title = title
        self.project_name = project_name
        self.date_range = date_range
        self.currency = currency_symbol
        self.firebase_client = firebase_client
        self.proyecto_id = proyecto_id

        if data is not None:
            raw_df = pd.DataFrame([dict(row) for row in data])
            if column_map and not raw_df.empty:
                cols_a_usar = [col for col in column_map.keys() if col in raw_df.columns]
                self.df = raw_df[cols_a_usar]
                self.df = self.df.rename(columns=column_map)
            else:
                self.df = raw_df
        else: 
            self.df = pd. DataFrame()

    def _clean_text_for_pdf(self, text:  str) -> str:
        """
        Limpia texto para PDF eliminando emojis y caracteres no soportados por latin-1.
        
        Args:
            text:  Texto a limpiar
            
        Returns:
            Texto limpio compatible con latin-1
        """
        if not isinstance(text, str):
            text = str(text)
        
        # Reemplazar emojis comunes por equivalentes ASCII
        emoji_map = {
            "📎": "[Adj]",
            "✅": "[OK]",
            "❌": "[X]",
            "⚠️": "[! ]",
            "🔗": "[Link]",
            "📄": "[Doc]",
            "📊": "[Chart]",
            "💰": "[$]",
            "🏦": "[Bank]",
            "📈": "[Up]",
            "📉": "[Down]",
        }
        
        for emoji, replacement in emoji_map. items():
            text = text. replace(emoji, replacement)
        
        # Intentar codificar como latin-1
        try:
            text. encode('latin-1')
            return text
        except UnicodeEncodeError:
            # Si falla, filtrar caracteres no-latin1
            cleaned = []
            for c in text:
                try: 
                    c.encode('latin-1')
                    cleaned.append(c)
                except UnicodeEncodeError:
                    cleaned.append('?')
            return ''.join(cleaned)


    def _crear_portada_anexo(self, trans:  dict, numero: int) -> str:
        """
        Crea una página de portada para un anexo. 
        
        Returns:
            Ruta al archivo PDF temporal creado
        """
        import tempfile
        
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='. pdf') as temp:
                temp_path = temp.name
            
            pdf = PDF(orientation='P', unit='mm', format='Letter')
            pdf.add_page()
            
            # Título del anexo
            pdf.set_font('Arial', 'B', 18)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(0, 15, f"ANEXO #{numero: 03d}", ln=True, align='C')
            pdf.ln(5)
            
            # Línea separadora
            pdf.set_draw_color(200, 200, 200)
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
            pdf.ln(10)
            
            # Tabla de detalles
            page_width = pdf.w - 2 * pdf.l_margin
            col1_width = page_width * 0.30
            col2_width = page_width * 0.70
            
            pdf.set_font('Arial', '', 10)
            pdf.set_text_color(0, 0, 0)
            
            # Fondo alternado
            row_bg_color = (245, 245, 245)
            
            detalles = [
                ("Fecha:", trans.get('fecha', '')),
                ("Descripción:", self._clean_text_for_pdf(trans.get('descripcion', '')[:80])),
                ("Categoría:", self._clean_text_for_pdf(trans.get('categoria', ''))),
                ("Subcategoría:", self._clean_text_for_pdf(trans.get('subcategoria', ''))),
                ("Monto:", f"{self.currency} {trans.get('monto', 0):,.2f}"),
                ("Adjuntos:", f"{len(trans.get('adjuntos_urls', []))} archivo(s)"),
            ]
            
            for idx, (label, value) in enumerate(detalles):
                if idx % 2 == 0:
                    pdf.set_fill_color(*row_bg_color)
                    fill = True
                else:
                    fill = False
                
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(col1_width, 8, label, border=1, fill=fill)
                
                pdf.set_font('Arial', '', 10)
                if label == "Monto: ":
                    pdf.set_text_color(192, 57, 43)  # Rojo para gastos
                else:
                    pdf.set_text_color(0, 0, 0)
                
                pdf.cell(col2_width, 8, str(value), border=1, ln=True, fill=fill)
            
            pdf.output(temp_path)
            return temp_path
        
        except Exception as e: 
            logger.error(f"Error creando portada de anexo:  {e}")
            return None


    def _descargar_adjunto_para_merger(self, adjunto_url: str) -> str:
        """
        Descarga un adjunto y devuelve la ruta local.
        Simplificado para usar con pypdf.
        
        Returns:
            Ruta al archivo local descargado, o None si falla
        """
        import tempfile
        import os
        from urllib.parse import urlparse, unquote
        
        try: 
            # Extraer path del Storage
            if 'firebasestorage. googleapis.com/v0/b/' in adjunto_url: 
                parsed = urlparse(adjunto_url)
                parts = parsed.path.split('/o/')
                if len(parts) >= 2:
                    storage_path = unquote(parts[1])
                else:
                    return None
            
            elif 'storage.googleapis.com' in adjunto_url:
                parsed = urlparse(adjunto_url)
                parts = parsed. path.split('/', 2)
                if len(parts) >= 3:
                    storage_path = unquote(parts[2]).split('? ')[0]
                else: 
                    return None
            
            elif adjunto_url.startswith('Proyecto/'):
                storage_path = adjunto_url
            
            else:
                logger.warning(f"Formato de URL no reconocido: {adjunto_url}")
                return None
            
            # Generar URL pública
            if not self. firebase_client: 
                logger.error("FirebaseClient no disponible")
                return None
            
            public_url = self.firebase_client.get_public_url_from_path(storage_path)
            
            # Obtener nombre de archivo
            filename = os.path.basename(storage_path)
            filename = filename.replace('? ', '_').replace('&', '_').replace('=', '_')
            
            # Descargar
            from progain4. utils. attachment_downloader import download_attachment
            
            local_file = download_attachment(public_url, filename)
            
            if local_file and os.path.exists(local_file):
                logger.debug(f"Descargado: {filename}")
                return local_file
            
            return None
        
        except Exception as e: 
            logger.error(f"Error descargando adjunto: {e}")
            return None


    def _imagen_a_pdf_temporal(self, imagen_path: str, titulo: str = "Anexo") -> str:
        """
        Convierte una imagen a PDF temporal.
        
        Args:
            imagen_path:  Ruta a la imagen
            titulo: Título para mostrar en el PDF
        
        Returns:
            Ruta al PDF temporal creado
        """
        import tempfile
        from PIL import Image
        
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp:
                temp_path = temp.name
            
            pdf = PDF(orientation='P', unit='mm', format='Letter')
            pdf.add_page()
            
            # Título
            pdf.set_font('Arial', 'B', 12)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(0, 10, titulo, ln=True, align='C')
            pdf.ln(5)
            
            # Calcular dimensiones
            with Image.open(imagen_path) as img:
                img_width_px, img_height_px = img. size
            
            # Convertir px a mm (96 DPI)
            px_to_mm = 25.4 / 96
            img_width_mm = img_width_px * px_to_mm
            img_height_mm = img_height_px * px_to_mm
            
            # Ajustar a página
            max_width = pdf.w - 30
            max_height = pdf. h - 60
            
            scale_w = max_width / img_width_mm if img_width_mm > max_width else 1
            scale_h = max_height / img_height_mm if img_height_mm > max_height else 1
            scale = min(scale_w, scale_h)
            
            final_width = img_width_mm * scale
            final_height = img_height_mm * scale
            
            # Centrar
            x_pos = (pdf.w - final_width) / 2
            y_pos = pdf.get_y()
            
            # Insertar imagen
            pdf.image(imagen_path, x=x_pos, y=y_pos, w=final_width)
            
            pdf.output(temp_path)
            return temp_path
        
        except Exception as e: 
            logger.error(f"Error convirtiendo imagen a PDF: {e}")
            return None

    def _crear_tabla_contenidos(self, transacciones:  list) -> str:
        """
        Crea una tabla de contenidos con enlaces clickeables.
        
        Returns:
            Ruta al PDF temporal de la tabla de contenidos
        """
        import tempfile
        from fpdf import FPDF
        
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp:
                temp_path = temp. name
            
            pdf = FPDF(orientation='P', unit='mm', format='Letter')
            pdf.add_page()
            
            # Título
            pdf.set_font('Arial', 'B', 18)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(0, 15, "TABLA DE CONTENIDOS - ANEXOS", ln=True, align='C')
            pdf.ln(5)
            
            # Línea separadora
            pdf.set_draw_color(200, 200, 200)
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
            pdf.ln(8)
            
            # Tabla
            pdf.set_font('Arial', '', 10)
            pdf.set_text_color(0, 0, 0)
            
            page_width = pdf.w - 2 * pdf.l_margin
            col_widths = [15, 25, 95, 45]
            
            # Header
            pdf.set_fill_color(44, 62, 80)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Arial', 'B', 9)
            
            headers = ["#", "Fecha", "Descripción", "Monto"]
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
            pdf.ln()
            
            # Filas
            pdf.set_font('Arial', '', 9)
            pdf.set_text_color(0, 0, 0)
            
            for idx, trans in enumerate(transacciones, 1):
                # Alternar color de fondo
                if idx % 2 == 0:
                    pdf.set_fill_color(245, 245, 245)
                    fill = True
                else:
                    fill = False
                
                # Número
                pdf.cell(col_widths[0], 7, str(idx), border=1, align='C', fill=fill)
                
                # Fecha
                fecha = trans.get('fecha', '')
                pdf.cell(col_widths[1], 7, fecha, border=1, align='C', fill=fill)
                
                # Descripción
                descripcion = self._clean_text_for_pdf(trans.get('descripcion', '')[:40])
                pdf.cell(col_widths[2], 7, descripcion, border=1, align='L', fill=fill)
                
                # Monto
                monto = trans.get('monto', 0)
                pdf.set_text_color(192, 57, 43)
                pdf.cell(col_widths[3], 7, f"{self.currency} {monto:,.2f}", border=1, align='R', fill=fill)
                pdf.set_text_color(0, 0, 0)
                
                pdf.ln()
            
            # Footer
            pdf.ln(5)
            pdf.set_font('Arial', 'I', 8)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(0, 5, f"Total de anexos: {len(transacciones)}", align='C')
            
            pdf.output(temp_path)
            return temp_path
        
        except Exception as e: 
            logger.error(f"Error creando tabla de contenidos: {e}")
            return None


    def _crear_portada_anexo_con_miniatura(self, trans:  dict, numero:  int) -> str:
        """
        Crea portada de anexo con miniatura del primer adjunto.
        
        Returns:
            Ruta al PDF temporal
        """
        import tempfile
        from fpdf import FPDF
        from PIL import Image
        import os
        
        try: 
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp:
                temp_path = temp.name
            
            pdf = FPDF(orientation='P', unit='mm', format='Letter')
            pdf.add_page()
            
            # Título
            pdf.set_font('Arial', 'B', 18)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(0, 15, f"ANEXO #{numero: 03d}", ln=True, align='C')
            pdf.ln(3)
            
            # Línea separadora
            pdf.set_draw_color(200, 200, 200)
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
            pdf.ln(8)
            
            # ========== MINIATURA ==========
            adjuntos_urls = trans.get('adjuntos_urls', [trans.get('adjunto_url')])
            
            if adjuntos_urls and adjuntos_urls[0]: 
                try:
                    # Descargar primer adjunto
                    local_file = self._descargar_adjunto_para_merger(adjuntos_urls[0])
                    
                    if local_file:
                        ext = os.path.splitext(local_file)[1].lower()
                        
                        # Solo mostrar miniatura si es imagen
                        if ext in ['. jpg', '.jpeg', '.png', '.gif', '.bmp']: 
                            # Miniatura centrada
                            max_thumb_width = 60
                            max_thumb_height = 60
                            
                            with Image.open(local_file) as img:
                                thumb_w_px, thumb_h_px = img. size
                            
                            px_to_mm = 25.4 / 96
                            thumb_w_mm = thumb_w_px * px_to_mm
                            thumb_h_mm = thumb_h_px * px_to_mm
                            
                            scale_w = max_thumb_width / thumb_w_mm if thumb_w_mm > max_thumb_width else 1
                            scale_h = max_thumb_height / thumb_h_mm if thumb_h_mm > max_thumb_height else 1
                            scale = min(scale_w, scale_h)
                            
                            final_w = thumb_w_mm * scale
                            final_h = thumb_h_mm * scale
                            
                            # Centrar miniatura
                            x_thumb = (pdf.w - final_w) / 2
                            y_thumb = pdf.get_y()
                            
                            # Fondo gris claro
                            pdf.set_fill_color(245, 245, 245)
                            pdf.rect(x_thumb - 2, y_thumb - 2, final_w + 4, final_h + 4, 'F')
                            
                            pdf.image(local_file, x=x_thumb, y=y_thumb, w=final_w)
                            pdf.set_y(y_thumb + final_h + 6)
                            
                            # Limpieza
                            try:
                                os.unlink(local_file)
                            except:
                                pass
                        
                        elif ext == '.pdf':
                            # Ícono de PDF
                            pdf. set_font('Arial', 'B', 48)
                            pdf.set_text_color(220, 38, 38)
                            pdf.cell(0, 30, "PDF", ln=True, align='C')
                            pdf.ln(5)
                
                except Exception as e: 
                    logger.debug(f"No se pudo generar miniatura: {e}")
            
            # ========== DETALLES ==========
            pdf.ln(5)
            page_width = pdf.w - 2 * pdf.l_margin
            col1_width = page_width * 0.30
            col2_width = page_width * 0.70
            
            pdf.set_font('Arial', '', 10)
            pdf.set_text_color(0, 0, 0)
            
            row_bg_color = (245, 245, 245)
            
            detalles = [
                ("Fecha:", trans.get('fecha', '')),
                ("Descripción:", self._clean_text_for_pdf(trans.get('descripcion', '')[:80])),
                ("Categoría:", self._clean_text_for_pdf(trans.get('categoria', ''))),
                ("Subcategoría:", self._clean_text_for_pdf(trans.get('subcategoria', ''))),
                ("Monto:", f"{self.currency} {trans.get('monto', 0):,.2f}"),
                ("Adjuntos:", f"{len(adjuntos_urls)} archivo(s)"),
            ]
            
            for idx, (label, value) in enumerate(detalles):
                if idx % 2 == 0:
                    pdf.set_fill_color(*row_bg_color)
                    fill = True
                else:
                    fill = False
                
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(col1_width, 8, label, border=1, fill=fill)
                
                pdf.set_font('Arial', '', 10)
                if label == "Monto: ":
                    pdf.set_text_color(192, 57, 43)
                else:
                    pdf.set_text_color(0, 0, 0)
                
                pdf.cell(col2_width, 8, str(value), border=1, ln=True, fill=fill)
            
            pdf.output(temp_path)
            return temp_path
        
        except Exception as e: 
            logger.error(f"Error creando portada con miniatura: {e}")
            return None

    def _procesar_adjunto_para_pdf(self, pdf, adjunto_url:   str, max_width: float, max_height: float) -> bool:
        """
        Descarga y procesa un adjunto para incluirlo en el PDF.
        
        ✅ CORREGIDO: Maneja URLs de Storage y genera URLs públicas correctas.  
        ✅ NUEVO: Busca Poppler automáticamente en ubicaciones comunes de Windows.
        
        Args:
            pdf:   Instancia de FPDF
            adjunto_url:  URL del adjunto en Firebase Storage
            max_width:   Ancho máximo en mm
            max_height:   Altura máxima en mm
        
        Returns:
            True si se procesó exitosamente, False en caso contrario
        """
        import tempfile
        import os
        from urllib.parse import urlparse, unquote
        
        try:
            # ✅ PASO 1: Extraer path limpio de la URL
            logger.debug(f"Procesando adjunto: {adjunto_url[: 100]}...")
            
            # Parsear URL
            parsed = urlparse(adjunto_url)
            
            # Extraer path del formato: https://storage.googleapis.com/BUCKET/PATH
            # o https://firebasestorage.googleapis.com/v0/b/BUCKET/o/ENCODED_PATH? alt=media
            
            if 'firebasestorage.googleapis.com/v0/b/' in adjunto_url:
                # Formato API:   /v0/b/BUCKET/o/ENCODED_PATH?alt=media
                # Extraer de /o/ hasta ?  
                parts = parsed.path.split('/o/')
                if len(parts) >= 2:
                    # Obtener ENCODED_PATH y decodificar
                    encoded_path = parts[1]
                    storage_path = unquote(encoded_path)
                    logger.debug(f"Extraído path (formato API): {storage_path}")
                else:
                    logger.warning(f"No se pudo extraer path de URL API: {adjunto_url}")
                    return False
            
            elif 'storage.googleapis.com' in adjunto_url:
                # Formato directo: https://storage.googleapis.com/BUCKET/PATH
                # Path empieza después de /BUCKET/
                parts = parsed.path.split('/', 2)  # ['', 'BUCKET', 'PATH']
                if len(parts) >= 3:
                    storage_path = unquote(parts[2])
                    # Quitar parámetros de query si existen (Expires, GoogleAccessId, etc.)
                    storage_path = storage_path.split('?')[0]
                    logger.debug(f"Extraído path (formato directo): {storage_path}")
                else:
                    logger.warning(f"No se pudo extraer path de URL directa: {adjunto_url}")
                    return False
            
            elif adjunto_url.startswith('Proyecto/') or adjunto_url.startswith('proyecto/'):
                # Ya es un path, usarlo directamente
                storage_path = adjunto_url
                logger.debug(f"Path directo detectado: {storage_path}")
            
            else:
                logger.warning(f"Formato de URL no reconocido:   {adjunto_url}")
                return False
            
            # ✅ PASO 2: Generar URL pública válida
            if not self. firebase_client: 
                logger.error("FirebaseClient no disponible")
                return False
            
            try:
                public_url = self.firebase_client.get_public_url_from_path(storage_path)
                logger.debug(f"URL pública:   {public_url[: 80]}...")
            except Exception as e:
                logger.error(f"Error generando URL pública:  {e}")
                return False
            
            # ✅ PASO 3: Obtener nombre de archivo limpio
            filename = os.path.basename(storage_path)
            # Limpiar caracteres problemáticos de Windows
            filename = filename.replace('? ', '_').replace('&', '_').replace('=', '_')
            
            logger.debug(f"Descargando como: {filename}")
            
            # ✅ PASO 4: Descargar archivo
            from progain4. utils. attachment_downloader import download_attachment
            
            local_file = download_attachment(public_url, filename)
            if not local_file or not os.path.exists(local_file):
                logger.warning(f"Descarga falló: {filename}")
                return False
            
            logger.info(f"✅ Descargado:   {filename} ({os.path.getsize(local_file)} bytes)")
            
            ext = os.path.splitext(filename)[1].lower()
            
            # === PROCESAR IMAGEN ===
            if ext in ['.jpg', '.jpeg', '. png', '.gif', '.bmp']: 
                try:
                    from PIL import Image
                    
                    img = Image.open(local_file)
                    img_width_px, img_height_px = img. size
                    
                    # Convertir px a mm
                    px_to_mm = 25.4 / 96
                    img_width_mm = img_width_px * px_to_mm
                    img_height_mm = img_height_px * px_to_mm
                    
                    # Calcular escala
                    scale_w = max_width / img_width_mm if img_width_mm > max_width else 1
                    scale_h = max_height / img_height_mm if img_height_mm > max_height else 1
                    scale = min(scale_w, scale_h)
                    
                    final_width = img_width_mm * scale
                    final_height = img_height_mm * scale
                    
                    # Verificar espacio
                    if pdf.get_y() + final_height + 10 > (pdf.h - pdf.b_margin):
                        pdf. add_page()
                    
                    # Centrar
                    x_pos = pdf.l_margin + (max_width - final_width) / 2
                    
                    # Insertar
                    pdf.image(local_file, x=x_pos, y=pdf.get_y(), w=final_width)
                    pdf.ln(final_height + 3)
                    
                    # Limpiar
                    try:
                        os.unlink(local_file)
                    except:
                        pass
                    
                    logger.info(f"✅ Imagen agregada al PDF:  {filename}")
                    return True
                    
                except Exception as e:  
                    logger.error(f"Error procesando imagen {filename}:  {e}")
                    try: 
                        os.unlink(local_file)
                    except: 
                        pass
                    return False
            
            # === PROCESAR PDF ===
            elif ext == '.pdf':
                pdf2image_available = False
                try: 
                    from pdf2image import convert_from_path  # type: ignore[import]
                    pdf2image_available = True
                except ImportError:
                    logger.info("pdf2image no disponible - instala con: pip install pdf2image")
                
                if pdf2image_available:
                    try: 
                        # ✅ BUSCAR POPPLER AUTOMÁTICAMENTE EN WINDOWS
                        poppler_path = None
                        
                        # Lista de ubicaciones comunes donde puede estar Poppler
                        possible_poppler_paths = [
                            r"C:\poppler-25.12.0\Library\bin",
                            r"C:\Program Files\poppler-25.12.0\Library\bin",
                            r"C:\Program Files (x86)\poppler\Library\bin",
                            r"C:\poppler-25.12.0\Library\bin",
                            r"C:\Program Files\poppler-25.12.0\Library\bin",
                        ]
                        
                        for path in possible_poppler_paths: 
                            if os.path. exists(path):
                                poppler_path = path
                                logger.debug(f"Poppler encontrado en: {poppler_path}")
                                break
                        
                        if not poppler_path:
                            logger.debug("Poppler no encontrado en ubicaciones comunes, usando PATH del sistema")
                        
                        # Convertir PDF a imágenes (máximo 3 páginas)
                        images = convert_from_path(
                            local_file, 
                            dpi=150, 
                            first_page=1, 
                            last_page=3,
                            poppler_path=poppler_path  # None si debe usar PATH
                        )
                        
                        logger.info(f"PDF convertido a {len(images)} imagen(es)")
                        
                        for page_num, img in enumerate(images, 1):
                            # Guardar imagen temporal
                            with tempfile. NamedTemporaryFile(delete=False, suffix='.png') as temp_img:
                                img.save(temp_img. name, 'PNG')
                                temp_img_path = temp_img.name
                            
                            # Verificar espacio en página
                            if pdf.get_y() + max_height > (pdf.h - pdf.b_margin):
                                pdf.add_page()
                            
                            # Insertar imagen de la página del PDF
                            pdf.image(temp_img_path, x=pdf.l_margin, y=pdf.get_y(), w=max_width)
                            pdf.ln(max_height + 3)
                            
                            # Limpiar imagen temporal
                            try:
                                os.unlink(temp_img_path)
                            except: 
                                pass
                        
                        # Limpiar PDF original
                        try:
                            os.unlink(local_file)
                        except:
                            pass
                        
                        logger.info(f"✅ PDF convertido e insertado: {filename} ({len(images)} página(s))")
                        return True
                        
                    except Exception as e:
                        logger. error(f"Error convirtiendo PDF {filename}: {e}")
                        logger.info("Posibles causas:")
                        logger.info("  1. Poppler no está instalado")
                        logger. info("  2. Poppler no está en PATH ni en ubicaciones comunes")
                        logger.info("  Descarga Poppler desde: https://github.com/oschwartz10612/poppler-windows/releases")
                
                # Fallback:   solo mostrar enlace
                pdf. set_font('Arial', '', 9)
                pdf.set_text_color(100, 100, 100)
                clean_filename = self._clean_text_for_pdf(filename)
                pdf.cell(0, 6, f"[PDF] {clean_filename}", ln=True)
                
                if not pdf2image_available:
                    pdf.set_font('Arial', '', 8)
                    pdf.set_text_color(192, 57, 43)
                    pdf.cell(0, 5, "  (Instala pdf2image para incrustar:  pip install pdf2image)", ln=True)
                else:
                    pdf.set_font('Arial', '', 8)
                    pdf.set_text_color(192, 57, 43)
                    pdf.cell(0, 5, "  (Instala Poppler para convertir PDFs a imagenes)", ln=True)
                
                pdf.set_text_color(0, 0, 255)
                pdf.cell(0, 6, "  Ver en Firebase Storage", ln=True)
                pdf.set_text_color(0, 0, 0)
                pdf.ln(3)
                
                try:  
                    os.unlink(local_file)
                except: 
                    pass
                
                logger.info(f"ℹ️ PDF referenciado (no incrustado): {filename}")
                return True
            
            # === OTROS ARCHIVOS ===
            else: 
                pdf.set_font('Arial', '', 10)
                pdf.set_text_color(100, 100, 100)
                clean_filename = self._clean_text_for_pdf(filename)
                pdf.cell(0, 6, f"[Archivo] {clean_filename}", ln=True)
                pdf.set_text_color(0, 0, 255)
                pdf.cell(0, 6, "  Ver en Firebase Storage", ln=True)
                pdf.set_text_color(0, 0, 0)
                pdf.ln(3)
                
                try: 
                    os.unlink(local_file)
                except: 
                    pass
                
                logger. info(f"ℹ️ Archivo referenciado:  {filename}")
                return True
            
        except Exception as e: 
            logger.error(f"Error general procesando adjunto: {e}")
            import traceback
            logger.error(traceback. format_exc())
            return False



    def to_excel(self, filepath):
        if self.df.empty:
            return False, "No hay datos para exportar."
        try:
            df_export = self.df.copy()
            # Eliminar columnas internas si existen
            internal_cols = ["_raw_tipo", "_transaction_id", "_adjuntos_paths"]
            for col in internal_cols:
                if col in df_export.columns:
                    df_export = df_export.drop(columns=[col])
                
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                sheet_name = 'Reporte'
                df_export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)
                worksheet = writer.sheets[sheet_name]
                
                # Estilos
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                title_font = Font(bold=True, size=16)
                subtitle_font = Font(italic=True, size=12)
                
                # Títulos
                last_col_letter = get_column_letter(len(df_export.columns))
                worksheet. merge_cells(f'A1:{last_col_letter}1')
                worksheet['A1'] = self.title
                worksheet['A1'].font = title_font
                worksheet['A1'].alignment = Alignment(horizontal='center')
                
                worksheet.merge_cells(f'A2:{last_col_letter}2')
                worksheet['A2'] = f"Proyecto: {self.project_name} ({self.date_range})"
                worksheet['A2'].font = subtitle_font
                worksheet['A2'].alignment = Alignment(horizontal='center')
                
                # Encabezado de tabla
                for cell in worksheet[5]:
                    cell.font = header_font
                    cell. fill = header_fill
                
                # Ajustar anchos
                for i, col in enumerate(df_export.columns):
                    column_letter = get_column_letter(i+1)
                    max_len = max([len(str(val)) for val in df_export[col]] + [len(col)])
                    worksheet.column_dimensions[column_letter].width = min(max_len + 2, 50)

            return True, None
        except Exception as e: 
            return False, str(e)

    def to_pdf(self, filepath=None):
        """
        Genera un reporte PDF tabular genérico con soporte para adjuntos incrustados.
        """
        if self.df. empty:
            return False, "No hay datos para exportar."
        if not filepath:
            return False, "No se indicó archivo de destino."

        try:
            # 1. Configuración Inicial
            pdf = PDF(orientation='L', unit='mm', format='Letter', 
                    title=self.title, project_name=self.project_name, date_range=self.date_range)
            
            pdf.set_auto_page_break(auto=False)  # Control manual de saltos
            pdf.add_page()
            
            # Definir colores unificados (Paleta Moderna)
            COLOR_HEADER_BG = (44, 62, 80)      # Azul Oscuro
            COLOR_HEADER_TXT = (255, 255, 255)  # Blanco
            COLOR_ROW_ALT = (245, 245, 245)     # Gris muy claro para filas alternas
            
            # Colores para montos
            COLOR_INGRESO = (39, 174, 96)       # Verde Esmeralda
            COLOR_GASTO = (192, 57, 43)         # Rojo
            COLOR_NEUTRO = (44, 62, 80)         # Azul Oscuro (Texto normal)

            # Preparar columnas (excluir internas)
            cols_to_print = [c for c in self.df.columns if c not in ["_raw_tipo", "_transaction_id", "_adjuntos_paths"]]
            
            # Calcular anchos dinámicos
            page_width = pdf.w - 2 * pdf.l_margin
            
            # Pesos para distribución de ancho
            total_weight = 0
            weights = {}
            for col in cols_to_print:  
                c_lower = col.lower()
                if "descrip" in c_lower:   weights[col] = 3.5
                elif "cuenta" in c_lower or "categor" in c_lower:  weights[col] = 2
                elif "fecha" in c_lower: weights[col] = 1.2
                elif "monto" in c_lower or "balance" in c_lower: weights[col] = 1.5
                elif "tipo" in c_lower: weights[col] = 1
                elif "adjunto" in c_lower: weights[col] = 0.8
                else: weights[col] = 1.5
                total_weight += weights[col]
            
            col_widths = {col: (weights[col] / total_weight) * page_width for col in cols_to_print}
            line_height = 7

            # Función para imprimir encabezado de tabla
            def print_table_header():
                pdf.set_font('Arial', 'B', 10)
                pdf.set_fill_color(*COLOR_HEADER_BG)
                pdf.set_text_color(*COLOR_HEADER_TXT)
                x = pdf.l_margin
                for col in cols_to_print:  
                    w = col_widths[col]
                    # ✅ Limpiar texto del encabezado
                    clean_col = self._clean_text_for_pdf(str(col))
                    pdf.cell(w, 9, clean_col, border=0, align='C', fill=True)
                pdf.ln(9)
                # Restaurar colores base
                pdf.set_text_color(0, 0, 0)
                pdf.set_font('Arial', '', 9)

            print_table_header()

            # Iterar filas
            total_ingresos = 0.0
            total_gastos = 0.0
            fill = False  # Para alternar colores

            for idx, row in self. df.iterrows():
                # 1. Calcular altura requerida (multi_cell)
                max_lines = 1
                for col in cols_to_print: 
                    val = str(row[col])
                    w = col_widths[col]
                    # Estimación de líneas
                    txt_width = pdf.get_string_width(val)
                    if txt_width > (w - 4):  # Margen interno
                        lines = int(txt_width / (w - 4)) + 1
                        if lines > max_lines:  max_lines = lines
                
                row_height = max_lines * line_height

                # 2. Salto de página si no cabe
                if pdf.get_y() + row_height > (pdf.h - pdf.b_margin):
                    pdf.add_page()
                    print_table_header()
                    fill = False

                # 3. Dibujar fondo alterno (Zebra striping)
                x_start = pdf.l_margin
                y_start = pdf.get_y()
                
                if fill: 
                    pdf.set_fill_color(*COLOR_ROW_ALT)
                    # Dibujar rectángulo de fondo para toda la fila
                    pdf.rect(x_start, y_start, page_width, row_height, 'F')
                
                # 4. Escribir celdas
                # Detectar tipo para colorear montos
                tipo_val = ""
                if "_raw_tipo" in self.df.columns:
                    tipo_val = str(row["_raw_tipo"]).lower()
                elif "Tipo" in row:  
                    tipo_val = str(row["Tipo"]).lower()

                x_curr = x_start
                for col in cols_to_print: 
                    val = str(row[col])
                    w = col_widths[col]
                    
                    # Estilo por columna
                    align = 'L'
                    text_rgb = COLOR_NEUTRO
                    
                    col_lower = col.lower()
                    
                    # Lógica de colores para Montos/Balance
                    if "monto" in col_lower or "balance" in col_lower or "ingresos" in col_lower or "gastos" in col_lower:
                        align = 'R'
                        try:
                            # Limpiar símbolos si vienen pre-formateados para detectar signo
                            clean_val = str(val).replace(self. currency, "").replace(",", "").strip()
                            num_val = float(clean_val)
                            
                            # Si es columna específica (ej.  Ingresos)
                            if "ingreso" in col_lower:  text_rgb = COLOR_INGRESO
                            elif "gasto" in col_lower:  text_rgb = COLOR_GASTO
                            elif "balance" in col_lower: 
                                text_rgb = COLOR_INGRESO if num_val >= 0 else COLOR_GASTO
                            # Si es columna genérica "Monto", dependemos del Tipo
                            elif "monto" in col_lower:  
                                if "ingreso" in tipo_val: text_rgb = COLOR_INGRESO
                                elif "gasto" in tipo_val: text_rgb = COLOR_GASTO
                                
                                # Acumular totales globales
                                if "ingreso" in tipo_val:  total_ingresos += abs(num_val)
                                elif "gasto" in tipo_val: total_gastos += abs(num_val)
                                
                            # Formatear bonito si es número puro
                            val = f"{self.currency} {num_val: ,.2f}"
                        except:  pass
                        
                    elif "tipo" in col_lower:  
                        align = 'C'
                        if "ingreso" in val.lower(): text_rgb = COLOR_INGRESO
                        elif "gasto" in val.lower(): text_rgb = COLOR_GASTO

                    pdf.set_xy(x_curr, y_start)
                    pdf.set_text_color(*text_rgb)
                    
                    # ✅ Limpiar texto antes de escribir
                    clean_val = self._clean_text_for_pdf(val)
                    
                    # pdf.cell no soporta multiline con height automático, usamos multi_cell
                    pdf.multi_cell(w, line_height, clean_val, border=0, align=align)
                    
                    x_curr += w

                # Siguiente fila
                pdf.set_y(y_start + row_height)
                fill = not fill  # Alternar color

            # Totales Finales
            pdf.ln(5)
            # Verificar espacio para totales
            if pdf.get_y() + 30 > (pdf.h - pdf.b_margin):
                pdf.add_page()
            
            # Dibujar caja de totales
            pdf.set_fill_color(240, 240, 240)
            pdf.rect(pdf.l_margin, pdf.get_y(), page_width, 20, 'F')
            
            pdf.set_y(pdf.get_y() + 5)
            pdf.set_font('Arial', 'B', 11)
            pdf.set_text_color(*COLOR_NEUTRO)
            
            col_w = page_width / 3
            
            # Ingresos
            pdf.set_x(pdf.l_margin)
            pdf.cell(col_w, 10, "Total Ingresos", 0, 0, 'C')
            
            # Gastos
            pdf.cell(col_w, 10, "Total Gastos", 0, 0, 'C')
            
            # Balance
            pdf.cell(col_w, 10, "Balance Neto", 0, 1, 'C')
            
            # Valores
            pdf.set_font('Arial', 'B', 12)
            pdf.ln(8)
            
            pdf.set_x(pdf.l_margin)
            pdf.set_text_color(*COLOR_INGRESO)
            pdf.cell(col_w, 10, f"{self.currency} {total_ingresos:,.2f}", 0, 0, 'C')
            
            pdf.set_text_color(*COLOR_GASTO)
            pdf.cell(col_w, 10, f"{self.currency} {total_gastos:,.2f}", 0, 0, 'C')
            
            balance = total_ingresos - total_gastos
            color_bal = COLOR_INGRESO if balance >= 0 else COLOR_GASTO
            pdf.set_text_color(*color_bal)
            pdf.cell(col_w, 10, f"{self.currency} {balance:,. 2f}", 0, 1, 'C')

            # ========== SECCIÓN DE ADJUNTOS ==========
            if self. firebase_client and self.proyecto_id and "_adjuntos_paths" in self.df. columns:
                # Recolectar transacciones con adjuntos
                transacciones_con_adjuntos = []
                for idx, row in self.df.iterrows():
                    adjuntos_paths = row. get("_adjuntos_paths", [])
                    if adjuntos_paths and len(adjuntos_paths) > 0:
                        transacciones_con_adjuntos.append({
                            "id": row. get("_transaction_id", ""),
                            "fecha": row.get("Fecha", ""),
                            "descripcion": row.get("Descripción", ""),
                            "adjuntos_paths": adjuntos_paths
                        })
                
                if transacciones_con_adjuntos:
                    # ✅ Importaciones con manejo de errores
                    try:
                        from progain4. utils.attachment_downloader import download_attachment
                    except ImportError:
                        logger.warning("attachment_downloader no disponible")
                        download_attachment = None
                    
                    try:
                        from PIL import Image
                    except ImportError:
                        logger. warning("PIL no disponible")
                        Image = None
                    
                    import tempfile
                    
                    logger.info(f"Processing {len(transacciones_con_adjuntos)} transactions with attachments")
                    
                    # Nueva página para adjuntos
                    pdf. add_page()
                    pdf.set_font('Arial', 'B', 16)
                    pdf.set_text_color(44, 62, 80)
                    pdf.cell(0, 10, 'ADJUNTOS DEL REPORTE', ln=True, align='C')
                    pdf.ln(5)
                    
                    # Procesar cada transacción
                    for trans in transacciones_con_adjuntos:
                        # Encabezado de transacción
                        pdf.set_font('Arial', 'B', 12)
                        pdf.set_text_color(44, 62, 80)
                        clean_header = self._clean_text_for_pdf(f"Transaccion: {trans['fecha']} - {trans['descripcion']}")
                        pdf.cell(0, 8, clean_header, ln=True)
                        pdf.ln(2)
                        
                        # Procesar cada adjunto
                        for path in trans['adjuntos_paths']:
                            try:
                                # Obtener URL pública
                                url = self.firebase_client.get_public_url_from_path(path)
                                filename = os.path.basename(path)
                                
                                # Descargar archivo
                                if not download_attachment:
                                    pdf.set_font('Arial', '', 10)
                                    pdf. set_text_color(192, 57, 43)
                                    pdf.cell(0, 6, "[X] Módulo de descarga no disponible", ln=True)
                                    continue
                                
                                local_file = download_attachment(url, filename)
                                if not local_file:
                                    pdf.set_font('Arial', '', 10)
                                    pdf. set_text_color(192, 57, 43)
                                    clean_error = self._clean_text_for_pdf(f"[X] Error descargando:  {filename}")
                                    pdf. cell(0, 6, clean_error, ln=True)
                                    continue
                                
                                ext = os.path.splitext(filename)[1].lower()
                                
                                # === PROCESAR PDF ===
                                if ext == '.pdf':
                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.set_text_color(39, 174, 96)
                                    clean_pdf_label = self._clean_text_for_pdf(f"[Adj] {filename} (PDF)")
                                    pdf.cell(0, 6, clean_pdf_label, ln=True)
                                    pdf.ln(2)
                                    
                                    # ✅ Intentar importar pdf2image de forma segura
                                    pdf2image_available = False
                                    try:
                                        from pdf2image import convert_from_path  # type: ignore[import]
                                        pdf2image_available = True
                                    except ImportError:
                                        pass
                                    
                                    if pdf2image_available:
                                        try:
                                            images = convert_from_path(local_file, dpi=150, first_page=1, last_page=3)
                                            
                                            for page_num, img in enumerate(images):
                                                # Guardar imagen temporal
                                                temp_img = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                                                img.save(temp_img.name, 'PNG')
                                                temp_img.close()
                                                
                                                # Agregar al PDF
                                                pdf.add_page()
                                                pdf.image(temp_img.name, x=10, y=30, w=pdf.w-20)
                                                
                                                # Limpiar
                                                try:
                                                    os.unlink(temp_img.name)
                                                except:
                                                    pass
                                        
                                        except Exception as e: 
                                            logger.error(f"Error convirtiendo PDF {filename}: {e}")
                                            pdf.set_font('Arial', '', 9)
                                            pdf.set_text_color(192, 57, 43)
                                            pdf.cell(0, 5, "  Error procesando PDF", ln=True)
                                    else:
                                        # Mostrar enlace si pdf2image no está disponible
                                        pdf.set_font('Arial', '', 9)
                                        pdf.set_text_color(100, 100, 100)
                                        pdf.cell(0, 5, "  (pdf2image no instalado)", ln=True)
                                        pdf.set_text_color(0, 0, 255)
                                        clean_url = self._clean_text_for_pdf(f"  Ver en linea: {url}")
                                        pdf.cell(0, 5, clean_url, ln=True, link=url)
                                    
                                    pdf.set_text_color(0, 0, 0)
                                
                                # === PROCESAR IMAGEN ===
                                elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                                    pdf.set_font('Arial', 'B', 10)
                                    pdf.set_text_color(39, 174, 96)
                                    clean_img_label = self._clean_text_for_pdf(f"[Adj] {filename}")
                                    pdf.cell(0, 6, clean_img_label, ln=True)
                                    pdf.ln(2)
                                    
                                    if Image: 
                                        try:
                                            # Obtener dimensiones de la imagen
                                            img = Image.open(local_file)
                                            img_width, img_height = img.size
                                            
                                            # Calcular dimensiones para ajustar a la página
                                            max_width_mm = pdf.w - 30
                                            max_height_mm = 160
                                            
                                            # Convertir píxeles a mm
                                            img_width_mm = img_width / 3.78
                                            img_height_mm = img_height / 3.78
                                            
                                            # Calcular escala
                                            scale_w = max_width_mm / img_width_mm if img_width_mm > max_width_mm else 1
                                            scale_h = max_height_mm / img_height_mm if img_height_mm > max_height_mm else 1
                                            scale = min(scale_w, scale_h)
                                            
                                            # Dimensiones finales
                                            final_width = img_width_mm * scale
                                            final_height = img_height_mm * scale
                                            
                                            # Verificar si cabe
                                            if pdf.get_y() + final_height + 10 > (pdf.h - pdf.b_margin):
                                                pdf.add_page()
                                            
                                            # Centrar imagen
                                            x_pos = (pdf.w - final_width) / 2
                                            
                                            # Agregar imagen
                                            pdf.image(local_file, x=x_pos, w=final_width)
                                            pdf.ln(5)
                                        
                                        except Exception as e:
                                            logger. error(f"Error procesando imagen {filename}: {e}")
                                            pdf.set_font('Arial', '', 9)
                                            pdf.set_text_color(192, 57, 43)
                                            pdf. cell(0, 5, "  Error procesando imagen", ln=True)
                                    else: 
                                        pdf.set_font('Arial', '', 9)
                                        pdf.set_text_color(192, 57, 43)
                                        pdf.cell(0, 5, "  PIL no disponible", ln=True)
                                
                                # === OTROS ARCHIVOS ===
                                else:
                                    pdf.set_font('Arial', '', 10)
                                    pdf.set_text_color(100, 100, 100)
                                    clean_file_label = self._clean_text_for_pdf(f"[Adj] {filename}")
                                    pdf.cell(0, 6, clean_file_label, ln=True)
                                    pdf.set_text_color(0, 0, 255)
                                    clean_link = self._clean_text_for_pdf(f"  [Link] Ver en linea: {url}")
                                    pdf.cell(0, 5, clean_link, ln=True, link=url)
                                    pdf.ln(2)
                                
                                # Limpiar archivo temporal
                                try: 
                                    os.unlink(local_file)
                                except: 
                                    pass
                            
                            except Exception as e: 
                                logger.error(f"Error procesando adjunto {path}: {e}")
                                pdf.set_font('Arial', '', 9)
                                pdf. set_text_color(192, 57, 43)
                                clean_error_msg = self._clean_text_for_pdf(f"[X] Error: {os.path.basename(path)}")
                                pdf.cell(0, 5, clean_error_msg, ln=True)
                        
                        pdf.ln(5)  # Espacio entre transacciones

            pdf.output(filepath)
            return True, None

        except Exception as e:
            logger. error(f"Error generating PDF: {e}", exc_info=True)
            return False, str(e)
    
    def to_excel_categoria(self, filepath):
        import pandas as pd
        from openpyxl. styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if self.df. empty:
            return False, "No hay datos para exportar."
        try:
            df = self.df. copy()
            df["Monto"] = pd.to_numeric(df["Monto"], errors='coerce').fillna(0)
            df_cats = df[(df["Subcategoría"]. isnull()) | (df["Subcategoría"] == "")]
            categorias = df_cats[df_cats["Categoría"] != "TOTAL GENERAL"]

            rows = []
            total_general = 0.0
            for _, cat_row in categorias.iterrows():
                cat = cat_row["Categoría"]
                total_categoria = cat_row["Monto"]
                rows.append({"Nivel": "Categoria", "Nombre": cat, "Monto":  total_categoria})
                total_general += total_categoria
                subcats = df[(df["Categoría"] == cat) & (df["Subcategoría"].notnull()) & (df["Subcategoría"] != "")]
                for _, sub_row in subcats.iterrows():
                    rows.append({"Nivel": "Subcategoria", "Nombre": sub_row["Subcategoría"], "Monto":  sub_row["Monto"]})
            rows.append({"Nivel": "Total", "Nombre": "TOTAL GENERAL", "Monto": total_general})

            df_export = pd.DataFrame(rows)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                sheet_name = 'Gastos por Categoría'
                df_export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)
                worksheet = writer.sheets[sheet_name]

                # Estilos profesionales
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="3778BE", end_color="3778BE", fill_type="solid")
                cat_font = Font(bold=True, color="FFFFFF")
                cat_fill = PatternFill(start_color="3778BE", end_color="3778BE", fill_type="solid")
                subcat_font = Font(bold=False, color="222222")
                subcat_fill = PatternFill(start_color="EBF0F5", end_color="EBF0F5", fill_type="solid")
                total_font = Font(bold=True, color="FFFFFF")
                total_fill = PatternFill(start_color="3CAADC", end_color="3CAADC", fill_type="solid")
                title_font = Font(bold=True, size=16)
                subtitle_font = Font(italic=True, size=12)
                center_align = Alignment(horizontal='center')
                right_align = Alignment(horizontal='right')
                currency_format = f'"{self.currency}" #,##0.00'
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                # Encabezados de título
                worksheet.merge_cells('A1:C1')
                title_cell = worksheet['A1']
                title_cell.value = "Reporte Profesional - Gastos por Categoría"
                title_cell.font = title_font
                title_cell.alignment = center_align
                worksheet.merge_cells('A2:C2')
                project_cell = worksheet['A2']
                project_cell.value = f"Proyecto: {self.project_name} ({self.date_range})"
                project_cell.font = subtitle_font
                project_cell.alignment = center_align

                # Encabezado de tabla
                header_row = worksheet[6]
                for cell in header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # Ajuste de ancho de columna
                for col_idx, col_name in enumerate(df_export.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = max(len(str(cell.value)) for cell in worksheet[column_letter] if cell.value)
                    worksheet. column_dimensions[column_letter]. width = max(max_length, 15) + 2

                # Formato de filas
                for idx, row in enumerate(df_export.itertuples(), start=7):
                    nivel = row.Nivel
                    nombre = row.Nombre
                    cell_nivel = worksheet[f"A{idx}"]
                    cell_nombre = worksheet[f"B{idx}"]
                    cell_monto = worksheet[f"C{idx}"]

                    cell_nombre.border = thin_border
                    cell_monto.border = thin_border
                    cell_nombre.alignment = Alignment(horizontal='left')
                    cell_monto.alignment = right_align
                    cell_monto.number_format = currency_format

                    if nivel == "Categoria":
                        cell_nombre.font = cat_font
                        cell_nombre.fill = cat_fill
                        cell_monto.font = cat_font
                        cell_monto.fill = cat_fill
                    elif nivel == "Subcategoria":
                        cell_nombre.font = subcat_font
                        cell_nombre. fill = subcat_fill
                        cell_monto.font = subcat_font
                        cell_monto.fill = subcat_fill
                    elif nivel == "Total":
                        cell_nombre.font = total_font
                        cell_nombre. fill = total_fill
                        cell_monto.font = total_font
                        cell_monto.fill = total_fill
                        cell_nombre.alignment = center_align
                        cell_monto.alignment = right_align

            return True, None
        except Exception as e:
            return False, str(e)

    def to_excel_resumen_por_cuenta(self, filepath):
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if self. df.empty:
            return False, "No hay datos para exportar."
        try:
            df_export = self.df.copy()

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                sheet_name = 'Resumen por Cuenta'
                df_export. to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)
                worksheet = writer.sheets[sheet_name]

                # ---- Estilos ----
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                ingreso_font = Font(bold=False, color="008000")  # Verde
                gasto_font = Font(bold=False, color="B40000")    # Rojo
                balance_font = Font(bold=False, color="000070")  # Azul
                title_font = Font(bold=True, size=16)
                subtitle_font = Font(italic=True, size=12)
                center_align = Alignment(horizontal='center')
                right_align = Alignment(horizontal='right')
                currency_format = f'"{self.currency}" #,##0.00'
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                # Título y subtítulo
                worksheet.merge_cells('A1:D1')
                title_cell = worksheet['A1']
                title_cell.value = self.title
                title_cell.font = title_font
                title_cell.alignment = center_align

                worksheet.merge_cells('A2:D2')
                project_cell = worksheet['A2']
                project_cell.value = f"Proyecto: {self.project_name}   |   Período: {self.date_range}"
                project_cell.font = subtitle_font
                project_cell.alignment = center_align

                # Encabezados
                header_row = worksheet[5]
                for cell in header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # Ajuste de ancho de columna
                for col_idx, col_name in enumerate(df_export.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = max(len(str(cell.value)) for cell in worksheet[column_letter] if cell.value)
                    worksheet.column_dimensions[column_letter].width = max(max_length, 15) + 2

                # Formato de filas
                for idx, row in enumerate(df_export.itertuples(), start=6):
                    # Cuenta
                    cell_cuenta = worksheet[f"A{idx}"]
                    cell_cuenta.alignment = Alignment(horizontal='left')
                    cell_cuenta. border = thin_border

                    # Ingresos
                    cell_ingreso = worksheet[f"B{idx}"]
                    cell_ingreso.number_format = currency_format
                    cell_ingreso.font = ingreso_font
                    cell_ingreso.alignment = right_align
                    cell_ingreso.border = thin_border

                    # Gastos
                    cell_gasto = worksheet[f"C{idx}"]
                    cell_gasto.number_format = currency_format
                    cell_gasto.font = gasto_font
                    cell_gasto.alignment = right_align
                    cell_gasto.border = thin_border

                    # Balance
                    cell_balance = worksheet[f"D{idx}"]
                    cell_balance.number_format = currency_format
                    cell_balance. font = balance_font
                    cell_balance.alignment = right_align
                    cell_balance. border = thin_border

            return True, None
        except Exception as e:
            return False, str(e)

    def to_pdf_resumen_por_cuenta(self, filepath=None):
        """
        Alias para usar el to_pdf genérico nuevo, ya que ahora soporta el estilo moderno
        que necesita el resumen por cuenta.
        """
        return self.to_pdf(filepath)

    def to_pdf_gastos_por_categoria(self, filepath=None, transacciones_anexos=None, progress_callback=None):
        """
        Genera PDF de Gastos por Categoría con miniaturas, TOC y progreso.
        
        Args:
            filepath:  Ruta del archivo PDF a generar
            transacciones_anexos: Lista de transacciones con adjuntos
            progress_callback:  Función callback(step, status, detail) para actualizar progreso
        
        Returns:
            Tuple (success:  bool, error_message: str or None)
        """
        if self.df. empty:
            return False, "No hay datos."
        if not filepath:
            return False, "Falta ruta."
        
        import tempfile
        import os
        import pandas as pd
        
        temp_files = []
        
        def update_progress(step, status, detail=""):
            """Helper para actualizar progreso."""
            if progress_callback:
                progress_callback(step, status, detail)
        
        try:
            # Calcular pasos totales
            num_anexos = len(transacciones_anexos) if transacciones_anexos else 0
            total_steps = 3 + num_anexos  # Reporte + TOC + Anexos + Merge
            current_step = 0
            
            update_progress(current_step, "Generando reporte principal.. .", "")
            
            # ========== FASE 1: GENERAR REPORTE PRINCIPAL ==========
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_report:
                temp_report_path = temp_report.name
                temp_files.append(temp_report_path)
            
            from fpdf import FPDF
            
            pdf = FPDF(orientation='P', unit='mm', format='Letter')
            pdf.set_auto_page_break(auto=True, margin=18)
            pdf.add_page()
            page_width = pdf.w - 2 * pdf.l_margin

            # Colores
            COLOR_BG_CAT = (44, 62, 80)
            COLOR_BG_TOTAL = (39, 174, 96)
            COLOR_BG_SUB = (236, 240, 241)
            COLOR_FONT_CAT = (255, 255, 255)
            COLOR_FONT_TOTAL = (255, 255, 255)
            COLOR_FONT_SUB = (44, 62, 80)

            # Header
            pdf.set_font("Arial", "B", 16)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(page_width, 10, f"Gastos por Categoria - {self.project_name}", ln=True, align='C')
            pdf.set_font("Arial", "", 11)
            pdf.cell(page_width, 8, f"Periodo: {self.date_range}", ln=True, align='C')
            pdf.ln(5)

            df = self.df.copy()
            
            if "Monto" in df.columns:
                df["Monto"] = pd.to_numeric(df["Monto"], errors='coerce').fillna(0)

            df_cats = df[(df["Subcategoría"]. isnull()) | (df["Subcategoría"] == "") | (df["Subcategoría"] == None)]
            categorias = df_cats[df_cats["Categoría"] != "TOTAL GENERAL"]
            total_general = 0.0

            # Renderizar categorías
            for _, cat_row in categorias.iterrows():
                cat = cat_row["Categoría"]
                total_categoria = cat_row["Monto"]
                total_general += total_categoria

                pdf.set_fill_color(*COLOR_BG_CAT)
                pdf.set_text_color(*COLOR_FONT_CAT)
                pdf.set_font("Arial", "B", 12)
                
                clean_cat = self._clean_text_for_pdf(f" {cat}")
                pdf.cell(int(page_width * 0.65), 9, clean_cat, border=0, align='L', fill=True)
                pdf.cell(int(page_width * 0.35), 9, f"{self.currency} {total_categoria:,.2f} ", border=0, align='R', fill=True)
                pdf.ln()

                subcats = df[(df["Categoría"] == cat) & (df["Subcategoría"].notnull()) & (df["Subcategoría"] != "")]
                pdf.set_font("Arial", "", 10)
                for _, sub_row in subcats. iterrows():
                    pdf.set_fill_color(*COLOR_BG_SUB)
                    pdf.set_text_color(*COLOR_FONT_SUB)
                    
                    clean_sub = self._clean_text_for_pdf(f"    {sub_row['Subcategoría']}")
                    pdf.cell(int(page_width * 0.65), 8, clean_sub, border=0, align='L', fill=True)
                    pdf.cell(int(page_width * 0.35), 8, f"{self.currency} {sub_row['Monto']:,.2f} ", border=0, align='R', fill=True)
                    pdf.ln()
                
                pdf.ln(1)

            # Total General
            pdf.ln(5)
            pdf.set_font("Arial", "B", 14)
            pdf.set_fill_color(*COLOR_BG_TOTAL)
            pdf.set_text_color(*COLOR_FONT_TOTAL)
            pdf.cell(int(page_width * 0.65), 12, " TOTAL GENERAL", border=0, align='L', fill=True)
            pdf.cell(int(page_width * 0.35), 12, f"{self.currency} {total_general:,.2f} ", border=0, align='R', fill=True)
            
            pdf.output(temp_report_path)
            
            current_step += 1
            update_progress(current_step, "Reporte principal completado", "Preparando tabla de contenidos...")
            
            # ========== FASE 2: CREAR TABLA DE CONTENIDOS ==========
            toc_path = None
            if transacciones_anexos and len(transacciones_anexos) > 0:
                toc_path = self._crear_tabla_contenidos(transacciones_anexos)
                if toc_path:
                    temp_files.append(toc_path)
            
            current_step += 1
            update_progress(current_step, "Tabla de contenidos creada", "Procesando anexos...")
            
            # ========== FASE 3: AGREGAR ANEXOS CON MINIATURAS ==========
            from pypdf import PdfWriter, PdfReader
            
            merger = PdfWriter()
            merger.append(PdfReader(temp_report_path, strict=False))
            
            if toc_path:
                merger.append(PdfReader(toc_path, strict=False))
            
            if transacciones_anexos and len(transacciones_anexos) > 0:
                for idx, trans in enumerate(transacciones_anexos, 1):
                    current_step += 1
                    descripcion = trans.get('descripcion', '')[: 30]
                    update_progress(
                        current_step, 
                        f"Procesando anexo {idx}/{num_anexos}",
                        f"{descripcion}..."
                    )
                    
                    try:
                        # Crear portada con miniatura
                        portada_anexo = self._crear_portada_anexo_con_miniatura(trans, idx)
                        if portada_anexo:
                            temp_files.append(portada_anexo)
                            merger.append(PdfReader(portada_anexo, strict=False))
                        
                        # Procesar adjuntos
                        adjuntos_urls = trans.get('adjuntos_urls', [trans.get('adjunto_url')])
                        
                        for adj_idx, adjunto_url in enumerate(adjuntos_urls, 1):
                            try:
                                local_file = self._descargar_adjunto_para_merger(adjunto_url)
                                
                                if not local_file:
                                    continue
                                
                                temp_files.append(local_file)
                                ext = os.path.splitext(local_file)[1].lower()
                                
                                if ext == '.pdf':
                                    merger. append(PdfReader(local_file, strict=False))
                                    logger.info(f"✅ PDF anexo #{idx}-{adj_idx} agregado")
                                
                                elif ext in ['.jpg', '.jpeg', '. png', '.gif', '.bmp']:
                                    temp_img_pdf = self._imagen_a_pdf_temporal(
                                        local_file, 
                                        f"Anexo #{idx} - Adjunto {adj_idx}"
                                    )
                                    
                                    if temp_img_pdf:
                                        temp_files.append(temp_img_pdf)
                                        merger.append(PdfReader(temp_img_pdf, strict=False))
                                        logger.info(f"✅ Imagen anexo #{idx}-{adj_idx} agregada")
                            
                            except Exception as e: 
                                logger.error(f"Error procesando adjunto:  {e}")
                    
                    except Exception as e:
                        logger.error(f"Error procesando transacción: {e}")
            
            # ========== FASE 4: ESCRIBIR PDF FINAL ==========
            update_progress(total_steps - 1, "Finalizando PDF...", "Guardando archivo...")
            
            with open(filepath, 'wb') as f_out:
                merger.write(f_out)
            
            update_progress(total_steps, "¡Completado!", f"PDF guardado:  {os.path.basename(filepath)}")
            
            logger.info(f"✅ PDF generado exitosamente: {filepath}")
            return True, None
        
        except Exception as e:
            logger. exception(f"Error generando PDF: {e}")
            return False, str(e)
        
        finally:
            for temp_file in temp_files: 
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                except:
                    pass

                
    def dashboard_to_pdf(self, filepath, figures, order=None):
        """
        Exporta un dashboard completo a PDF:  
        1. Gráficos (imágenes) al principio.  
        2. Tabla de datos detallada a continuación (usando self.df).
        """
        try:
            if not figures or not isinstance(figures, dict):
                return False, "No hay figuras para exportar."

            # 1. Configuración Inicial del PDF
            pdf = PDF(orientation='L', unit='mm', format='A4',
                    title=self.title,
                    project_name=self.project_name,
                    date_range=self.date_range)
            
            # --- FASE 1: GRÁFICOS ---
            default_order = ['grafico_principal', 'ivg', 'gastos_cat', 'ingresos_cat', 'gastos_subcat']
            keys = order or [k for k in default_order if k in figures] or list(figures.keys())

            exported_count = 0

            for key in keys:
                fig = figures.get(key)
                if fig is None:  continue

                real_fig = fig
                try:
                    from matplotlib. figure import Figure
                    if not isinstance(fig, Figure) and hasattr(fig, "figure"):
                        real_fig = fig.figure
                    
                    # Adaptación para Plotly
                    if hasattr(real_fig, 'to_image'):
                        import io
                        img_bytes = real_fig.to_image(format="png", width=1200, height=700)
                        with io.BytesIO(img_bytes) as buf:
                            pdf.add_page()
                            pdf.image(buf, x=10, y=30, w=277)
                            exported_count += 1
                        continue
                except ImportError:  pass

                # Lógica Matplotlib estándar
                if hasattr(real_fig, 'savefig'):
                    with io.BytesIO() as buf:
                        try:
                            real_fig. savefig(buf, format='png', dpi=150, bbox_inches='tight')
                            buf.seek(0)
                            pdf.add_page()
                            pdf.image(buf, x=10, y=30, w=277)
                            exported_count += 1
                        except Exception as e:
                            print(f"Error renderizando gráfico {key}: {e}")
                            continue

            # --- FASE 2: TABLA DE DATOS ---
            if not self.df.empty:
                pdf.add_page()
                pdf.ln(5)
                pdf.set_font('Arial', 'B', 14)
                pdf.set_text_color(44, 62, 80)
                pdf.cell(0, 10, "Detalle de Datos", 0, 1, 'L')
                pdf.ln(2)

                # Configuración de Estilo
                COLOR_HEADER_BG = (44, 62, 80)
                COLOR_HEADER_TXT = (255, 255, 255)
                COLOR_ROW_ALT = (245, 245, 245)
                COLOR_INGRESO = (39, 174, 96)
                COLOR_GASTO = (192, 57, 43)
                COLOR_NEUTRO = (44, 62, 80)

                cols_to_print = [c for c in self.df.columns if c not in ["_raw_tipo", "_transaction_id", "_adjuntos_paths"]]
                
                # Calcular anchos
                page_width = pdf.w - 2 * pdf.l_margin
                total_weight = 0
                weights = {}
                for col in cols_to_print: 
                    c_lower = col.lower()
                    if "descrip" in c_lower:  weights[col] = 3.5
                    elif "cuenta" in c_lower or "categor" in c_lower: weights[col] = 2
                    elif "fecha" in c_lower: weights[col] = 1.2
                    # Damos peso estándar a columnas numéricas
                    elif any(k in c_lower for k in ["monto", "balance", "ingreso", "gasto", "total"]): weights[col] = 1.5
                    elif "tipo" in c_lower: weights[col] = 1
                    else: weights[col] = 1.5
                    total_weight += weights[col]
                
                col_widths = {col: (weights[col] / total_weight) * page_width for col in cols_to_print}
                line_height = 7

                # Encabezado Tabla
                pdf.set_font('Arial', 'B', 10)
                pdf.set_fill_color(*COLOR_HEADER_BG)
                pdf.set_text_color(*COLOR_HEADER_TXT)
                for col in cols_to_print:
                    w = col_widths[col]
                    # ✅ Limpiar texto
                    clean_col = self._clean_text_for_pdf(str(col))
                    pdf. cell(w, 9, clean_col, border=0, align='C', fill=True)
                pdf.ln(9)
                
                pdf.set_font('Arial', '', 9)
                fill = False

                # Filas
                for idx, row in self. df.iterrows():
                    # Calcular altura
                    max_lines = 1
                    for col in cols_to_print:
                        val = str(row[col])
                        w = col_widths[col]
                        txt_width = pdf.get_string_width(val)
                        if txt_width > (w - 4):
                            lines = int(txt_width / (w - 4)) + 1
                            if lines > max_lines:  max_lines = lines
                    
                    row_height = max_lines * line_height

                    # Salto de página
                    if pdf.get_y() + row_height > (pdf.h - pdf.b_margin):
                        pdf. add_page()
                        # Repetir encabezado
                        pdf.set_font('Arial', 'B', 10)
                        pdf.set_fill_color(*COLOR_HEADER_BG)
                        pdf.set_text_color(*COLOR_HEADER_TXT)
                        for col in cols_to_print:
                            w = col_widths[col]
                            clean_col = self._clean_text_for_pdf(str(col))
                            pdf.cell(w, 9, clean_col, border=0, align='C', fill=True)
                        pdf.ln(9)
                        pdf.set_font('Arial', '', 9)
                        fill = False

                    # Dibujar fila
                    x_curr = pdf.l_margin
                    y_start = pdf.get_y()
                    
                    if fill:
                        pdf.set_fill_color(*COLOR_ROW_ALT)
                        pdf.rect(x_curr, y_start, page_width, row_height, 'F')

                    # Detectar tipo general de la fila (si existe) para colorear montos genéricos
                    tipo_row = ""
                    if "_raw_tipo" in self.df.columns: tipo_row = str(row["_raw_tipo"]).lower()

                    for col in cols_to_print: 
                        val = str(row[col])
                        w = col_widths[col]
                        align = 'L'
                        text_rgb = COLOR_NEUTRO
                        col_lower = col.lower()

                        # Detectamos columnas numéricas usando singular y plural:  "gasto", "gastos", "total", etc.
                        is_numeric_col = any(k in col_lower for k in ["monto", "balance", "ingreso", "gasto", "total"])
                        
                        if is_numeric_col:
                            align = 'R'
                            try:
                                # Limpiar y convertir a número
                                clean_val = str(val).replace(self.currency, "").replace(",", "").strip()
                                num_val = float(clean_val)
                                
                                # Lógica de colores
                                if "ingreso" in col_lower: 
                                    text_rgb = COLOR_INGRESO
                                elif "gasto" in col_lower:
                                    text_rgb = COLOR_GASTO
                                elif "balance" in col_lower:
                                    text_rgb = COLOR_INGRESO if num_val >= 0 else COLOR_GASTO
                                # Si la columna es genérica (ej "Monto" o "Total") usamos el tipo de fila
                                else:
                                    if "ingreso" in tipo_row:  text_rgb = COLOR_INGRESO
                                    elif "gasto" in tipo_row: text_rgb = COLOR_GASTO
                                
                                val = f"{self.currency} {num_val:,.2f}"
                            except:
                                # Si falla la conversión (ej.  es texto), dejamos el color neutro
                                pass
                        
                        elif "tipo" in col_lower: 
                            align = 'C'
                            if "ingreso" in val. lower(): text_rgb = COLOR_INGRESO
                            elif "gasto" in val.lower(): text_rgb = COLOR_GASTO

                        pdf.set_xy(x_curr, y_start)
                        pdf.set_text_color(*text_rgb)
                        
                        # ✅ Limpiar texto
                        clean_val = self._clean_text_for_pdf(val)
                        pdf. multi_cell(w, line_height, clean_val, border=0, align=align)
                        x_curr += w

                    pdf.set_y(y_start + row_height)
                    fill = not fill

            pdf.output(filepath)
            return True, None

        except Exception as e:
            return False, str(e)
        
