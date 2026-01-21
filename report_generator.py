import pandas as pd
import re
import io
import os
from fpdf import FPDF
from fpdf.enums import XPos, YPos
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class PDF(FPDF):
    def __init__(self, orientation='L', unit='mm', format='Letter', title="", project_name="", date_range=""):
        super().__init__(orientation, unit, format)
        self.report_title = title
        self.project_name = project_name
        self.date_range = date_range

    def header(self):
        try:
            self.set_font('Segoe', 'B', 16)
        except Exception:
            self.set_font('Helvetica', 'B', 16)
        self.cell(0, 10, self.report_title, 0, 1, 'C')
        try:
            self.set_font('Segoe', 'I', 10)
        except Exception:
            self.set_font('Helvetica', 'I', 10)
        self.cell(0, 6, f"Proyecto: {self.project_name}", 0, 1, 'C')
        self.cell(0, 6, f"Período: {self.date_range}", 0, 1, 'C')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        try:
            self.set_font('Segoe', 'I', 8)
        except Exception:
            self.set_font('Helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')

class ReportGenerator:
    def __init__(self, data=None, title="", project_name="", date_range="", currency_symbol="RD$", column_map=None):
        self.title = title
        self.project_name = project_name
        self.date_range = date_range
        self.currency = currency_symbol

        if data is not None:
            raw_df = pd.DataFrame([dict(row) for row in data])
            if column_map and not raw_df.empty:
                cols_a_usar = [col for col in column_map.keys() if col in raw_df.columns]
                self.df = raw_df[cols_a_usar]
                self.df = self.df.rename(columns=column_map)
            else:
                self.df = raw_df
        else:
            self.df = pd.DataFrame()

    def to_excel(self, filepath):
        if self.df.empty:
            return False, "No hay datos para exportar."
        try:
            df_export = self.df.copy()
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                sheet_name = 'Reporte'
                df_export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)
                worksheet = writer.sheets[sheet_name]
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                title_font = Font(bold=True, size=16)
                subtitle_font = Font(italic=True, size=12)
                total_font = Font(bold=True)
                center_align = Alignment(horizontal='center')
                right_align = Alignment(horizontal='right')
                currency_format = f'"{self.currency}" #,##0.00'
                last_col_letter = get_column_letter(df_export.shape[1])
                worksheet.merge_cells(f'A1:{last_col_letter}1')
                title_cell = worksheet['A1']
                title_cell.value = self.title
                title_cell.font = title_font
                title_cell.alignment = center_align
                worksheet.merge_cells(f'A2:{last_col_letter}2')
                project_cell = worksheet['A2']
                project_cell.value = f"Proyecto: {self.project_name} ({self.date_range})"
                project_cell.font = subtitle_font
                project_cell.alignment = center_align
                header_row = worksheet[5]
                for cell in header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                numeric_cols_indices = [i for i, dtype in enumerate(df_export.dtypes) if pd.api.types.is_numeric_dtype(dtype)]
                for col_idx, col_name in enumerate(df_export.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = max(len(str(cell.value)) for cell in worksheet[column_letter] if cell.value)
                    worksheet.column_dimensions[column_letter].width = max_length + 2
                    if col_idx - 1 in numeric_cols_indices:
                        for cell in worksheet[column_letter][6:]:
                            cell.number_format = currency_format
                if numeric_cols_indices:
                    total_row_idx = len(df_export) + 6
                    for col_idx in numeric_cols_indices:
                        col_letter = get_column_letter(col_idx + 1)
                        sum_formula = f"=SUM({col_letter}6:{col_letter}{total_row_idx - 1})"
                        if col_idx == numeric_cols_indices[0]:
                            label_cell_col = get_column_letter(col_idx)
                            label_cell = worksheet[f"{label_cell_col}{total_row_idx}"]
                            label_cell.value = "Total:"
                            label_cell.font = total_font
                            label_cell.alignment = right_align
                        value_cell = worksheet[f"{col_letter}{total_row_idx}"]
                        value_cell.value = sum_formula
                        value_cell.number_format = currency_format
                        value_cell.font = total_font
            return True, None
        except Exception as e:
            return False, str(e)

    def to_pdf(self, filepath=None):
        import os
        import re
        if self.df.empty:
            return False, "No hay datos para exportar."
        try:
            if not filepath:
                return False, "No se indicó el nombre del archivo PDF."

            from fpdf import FPDF

            pdf = PDF(
                orientation='L',
                unit='mm',
                format='Letter',
                title=self.title,
                project_name=self.project_name,
                date_range=self.date_range
            )
            font_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts')
            segoe_regular = os.path.join(font_dir, 'seguisb.ttf')
            segoe_bold = os.path.join(font_dir, 'seguisb.ttf')
            segoe_italic = os.path.join(font_dir, 'seguisbi.ttf')
            roboto_regular = os.path.join(font_dir, 'Roboto-Regular.ttf')
            roboto_bold = os.path.join(font_dir, 'Roboto-Bold.ttf')
            roboto_italic = os.path.join(font_dir, 'Roboto-Italic.ttf')

            font_family = 'Helvetica'
            try:
                if os.path.exists(segoe_regular) and os.path.exists(segoe_bold) and os.path.exists(segoe_italic):
                    pdf.add_font('Segoe', '', segoe_regular, uni=True)
                    pdf.add_font('Segoe', 'B', segoe_bold, uni=True)
                    pdf.add_font('Segoe', 'I', segoe_italic, uni=True)
                    font_family = 'Segoe'
                elif os.path.exists(roboto_regular) and os.path.exists(roboto_bold) and os.path.exists(roboto_italic):
                    pdf.add_font('Roboto', '', roboto_regular, uni=True)
                    pdf.add_font('Roboto', 'B', roboto_bold, uni=True)
                    pdf.add_font('Roboto', 'I', roboto_italic, uni=True)
                    font_family = 'Roboto'
            except Exception:
                font_family = 'Helvetica'

            pdf.set_auto_page_break(auto=True, margin=18)
            pdf.add_page()

            # Elimina la columna "Tipo" para aprovechar espacio
            tipo_col_present = "Tipo" in self.df.columns
            if tipo_col_present:
                col_tipo = self.df["Tipo"].tolist()
                self.df = self.df.drop(columns=["Tipo"])
            else:
                col_tipo = [""] * len(self.df)

            col_names = self.df.columns.tolist()
            num_cols = len(col_names)
            page_width = pdf.w - 2 * pdf.l_margin

            # --- Calcular ancho dinámico de columnas ---
            col_max_lens = []
            for col in col_names:
                max_len = pdf.get_string_width(str(col))
                for val in self.df[col]:
                    max_len = max(max_len, pdf.get_string_width(str(val)))
                col_max_lens.append(max_len + 10)
            total_max_len = sum(col_max_lens)
            col_widths = [page_width * (w / total_max_len) for w in col_max_lens]

            # --- Encabezado de la tabla ---
            pdf.set_font(font_family, 'B', 11)
            pdf.set_fill_color(79, 129, 189)
            pdf.set_text_color(255, 255, 255)
            for i, header in enumerate(col_names):
                pdf.cell(col_widths[i], 10, header, border=1, align='C', fill=True)
            pdf.ln()

            # --- Filas de tabla ---
            pdf.set_font(font_family, '', 10)
            total_gastos = 0.0
            total_ingresos = 0.0

            for idx, row in self.df.iterrows():
                y_before = pdf.get_y()
                x_left = pdf.l_margin
                max_row_height = pdf.font_size + 6  # Default cell height

                # Pre-calcula el alto máximo de la fila por si hay texto largo
                for i, col in enumerate(col_names):
                    val = str(row[col])
                    nlines = max(1, (pdf.get_string_width(val) // (col_widths[i] - 2)))
                    cell_height = (pdf.font_size + 3) * nlines
                    max_row_height = max(max_row_height, cell_height)

                tipo = col_tipo[idx].lower() if tipo_col_present else ""
                for i, col in enumerate(col_names):
                    value = row[col]
                    align = 'L'
                    text_color = (0, 0, 0)
                    if col.lower() == "monto":
                        try:
                            monto_val = float(value)
                        except (ValueError, TypeError):
                            monto_val = 0.0
                        if "ingreso" in tipo:
                            value = f"{self.currency} {monto_val:,.2f}"
                            text_color = (0, 100, 0)
                            total_ingresos += monto_val
                        else:
                            value = f"{self.currency} -{monto_val:,.2f}"
                            text_color = (180, 0, 0)
                            total_gastos += monto_val
                        align = 'R'
                    else:
                        value = str(value)
                    pdf.set_text_color(*text_color)
                    pdf.set_xy(x_left, y_before)
                    pdf.multi_cell(col_widths[i], max_row_height, value, border=1, align=align)
                    x_left += col_widths[i]
                pdf.set_text_color(0, 0, 0)
                pdf.set_xy(pdf.l_margin, y_before + max_row_height)

            # --- Leyenda debajo de la tabla ---
            pdf.set_font(font_family, 'I', 9)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(page_width, 8, "Montos en verde: Ingresos   |   Montos en rojo y signo menos: Gastos.", 0, 1, 'L')

            # --- Totales en una tabla aparte ---
            pdf.ln(4)
            pdf.set_font(font_family, 'B', 11)
            total_table_width = page_width * 0.7
            col_total_width = total_table_width / 3

            # Encabezados de totales
            pdf.set_fill_color(79, 129, 189)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(col_total_width, 10, "Total Ingresos", border=1, align='C', fill=True)
            pdf.cell(col_total_width, 10, "Total Gastos", border=1, align='C', fill=True)
            pdf.cell(col_total_width, 10, "Balance", border=1, align='C', fill=True)
            pdf.ln()

            # Valores de totales
            pdf.set_font(font_family, 'B', 12)
            pdf.set_text_color(0, 100, 0)
            pdf.cell(col_total_width, 12, f"{self.currency} {total_ingresos:,.2f}", border=1, align='C')
            pdf.set_text_color(180, 0, 0)
            pdf.cell(col_total_width, 12, f"{self.currency} -{total_gastos:,.2f}", border=1, align='C')
            pdf.set_text_color(0, 0, 70)
            balance = total_ingresos - total_gastos
            pdf.cell(col_total_width, 12, f"{self.currency} {balance:,.2f}", border=1, align='C')
            pdf.set_text_color(0, 0, 0)
            pdf.ln()

            pdf.output(filepath)
            return True, None
        except Exception as e:
            return False, str(e)


    def to_pdf_gastos_por_categoria(self, filepath=None):
        import pandas as pd

        if self.df.empty:
            return False, "No hay datos para exportar."
        try:
            if not filepath:
                return False, "No se indicó el nombre del archivo PDF."
            from fpdf import FPDF

            pdf = FPDF(orientation='P', unit='mm', format='Letter')
            pdf.set_auto_page_break(auto=True, margin=18)
            pdf.add_page()

            page_width = pdf.w - 2 * pdf.l_margin

            # Colores modernos
            COLOR_BG_CAT = (44, 62, 80)
            COLOR_BG_TOTAL = (39, 174, 96)
            COLOR_BG_SUB = (236, 240, 241)
            COLOR_FONT_CAT = (255, 255, 255)
            COLOR_FONT_TOTAL = (255, 255, 255)
            COLOR_FONT_SUB = (44, 62, 80)

            # Títulos y encabezado
            pdf.set_font("Arial", "B", 16)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(page_width, 12, f"Gastos por Categoría - {self.project_name}", ln=True, align='L')
            pdf.set_font("Arial", "", 11)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(page_width, 8, f"Periodo: {self.date_range}", ln=True, align='L')
            pdf.ln(2)

            df = self.df.copy()
            df["Monto"] = pd.to_numeric(df["Monto"], errors='coerce').fillna(0)

            # Solo las filas de categorías (Subcategoría vacía) para total general
            df_cats = df[(df["Subcategoría"].isnull()) | (df["Subcategoría"] == "")]
            categorias = df_cats[df_cats["Categoría"] != "TOTAL GENERAL"]

            total_general = categorias["Monto"].sum()

            # --- Tabla moderna ---
            for _, cat_row in categorias.iterrows():
                cat = cat_row["Categoría"]
                total_categoria = cat_row["Monto"]
                # Categoría: fondo azul oscuro, texto blanco, negrita
                pdf.set_fill_color(*COLOR_BG_CAT)
                pdf.set_text_color(*COLOR_FONT_CAT)
                pdf.set_font("Arial", "B", 13)
                pdf.cell(int(page_width * 0.65), 9, f"{cat}", border=1, align='L', fill=True)
                pdf.cell(int(page_width * 0.35), 9, f"{self.currency} {total_categoria:,.2f}", border=1, align='R', fill=True)
                pdf.ln()

                # Subcategorías: fondo gris claro, texto azul oscuro
                subcats = df[(df["Categoría"] == cat) & (df["Subcategoría"].notnull()) & (df["Subcategoría"] != "")]
                pdf.set_font("Arial", "", 10)
                for _, sub_row in subcats.iterrows():
                    pdf.set_fill_color(*COLOR_BG_SUB)
                    pdf.set_text_color(*COLOR_FONT_SUB)
                    pdf.cell(int(page_width * 0.65), 8, f"    {sub_row['Subcategoría']}", border=1, align='L', fill=True)
                    pdf.cell(int(page_width * 0.35), 8, f"{self.currency} {sub_row['Monto']:,.2f}", border=1, align='R', fill=True)
                    pdf.ln()
                pdf.ln(1)

            # Total general al final
            pdf.set_font("Arial", "B", 14)
            pdf.set_fill_color(*COLOR_BG_TOTAL)
            pdf.set_text_color(*COLOR_FONT_TOTAL)
            pdf.cell(int(page_width * 0.65), 10, "TOTAL GENERAL", border=1, align='L', fill=True)
            pdf.cell(int(page_width * 0.35), 10, f"{self.currency} {total_general:,.2f}", border=1, align='R', fill=True)
            pdf.ln(15)

            pdf.set_text_color(100, 100, 100)
            pdf.set_font("Arial", "I", 9)
            pdf.cell(page_width, 8, "Las categorías están resaltadas en azul. Totales en verde. Subcategorías en gris.", 0, 1, 'L')

            pdf.output(filepath)
            return True, None
        except Exception as e:
            return False, str(e)


    def to_excel_categoria(self, filepath):
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if self.df.empty:
            return False, "No hay datos para exportar."
        try:
            df = self.df.copy()
            df["Monto"] = pd.to_numeric(df["Monto"], errors='coerce').fillna(0)
            df_cats = df[(df["Subcategoría"].isnull()) | (df["Subcategoría"] == "")]
            categorias = df_cats[df_cats["Categoría"] != "TOTAL GENERAL"]

            rows = []
            total_general = 0.0
            for _, cat_row in categorias.iterrows():
                cat = cat_row["Categoría"]
                total_categoria = cat_row["Monto"]
                rows.append({"Nivel": "Categoria", "Nombre": cat, "Monto": total_categoria})
                total_general += total_categoria
                subcats = df[(df["Categoría"] == cat) & (df["Subcategoría"].notnull()) & (df["Subcategoría"] != "")]
                for _, sub_row in subcats.iterrows():
                    rows.append({"Nivel": "Subcategoria", "Nombre": sub_row["Subcategoría"], "Monto": sub_row["Monto"]})
            rows.append({"Nivel": "Total", "Nombre": "TOTAL GENERAL", "Monto": total_general})

            df_export = pd.DataFrame(rows)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                sheet_name = 'Gastos por Categoría'
                df_export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)
                worksheet = writer.sheets[sheet_name]

                # Estilos profesionales
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="3778BE", end_color="3778BE", fill_type="solid")
                cat_font = Font(bold=True, color="FFFFFF")
                cat_fill = PatternFill(start_color="3778BE", end_color="3778BE", fill_type="solid")
                subcat_font = Font(bold=False, color="222222")
                subcat_fill = PatternFill(start_color="EBF0F5", end_color="EBF0F5", fill_type="solid")
                total_font = Font(bold=True, color="FFFFFF")
                total_fill = PatternFill(start_color="3CAADC", end_color="3CAADC", fill_type="solid")
                title_font = Font(bold=True, size=16)
                subtitle_font = Font(italic=True, size=12)
                center_align = Alignment(horizontal='center')
                right_align = Alignment(horizontal='right')
                currency_format = f'"{self.currency}" #,##0.00'
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                # Encabezados de título
                worksheet.merge_cells('A1:C1')
                title_cell = worksheet['A1']
                title_cell.value = "Reporte Profesional - Gastos por Categoría"
                title_cell.font = title_font
                title_cell.alignment = center_align
                worksheet.merge_cells('A2:C2')
                project_cell = worksheet['A2']
                project_cell.value = f"Proyecto: {self.project_name} ({self.date_range})"
                project_cell.font = subtitle_font
                project_cell.alignment = center_align

                # Encabezado de tabla
                header_row = worksheet[6]
                for cell in header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # Ajuste de ancho de columna
                for col_idx, col_name in enumerate(df_export.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = max(len(str(cell.value)) for cell in worksheet[column_letter] if cell.value)
                    worksheet.column_dimensions[column_letter].width = max(max_length, 15) + 2

                # Formato de filas
                for idx, row in enumerate(df_export.itertuples(), start=7):
                    nivel = row.Nivel
                    nombre = row.Nombre
                    cell_nivel = worksheet[f"A{idx}"]
                    cell_nombre = worksheet[f"B{idx}"]
                    cell_monto = worksheet[f"C{idx}"]

                    cell_nombre.border = thin_border
                    cell_monto.border = thin_border
                    cell_nombre.alignment = Alignment(horizontal='left')
                    cell_monto.alignment = right_align
                    cell_monto.number_format = currency_format

                    if nivel == "Categoria":
                        cell_nombre.font = cat_font
                        cell_nombre.fill = cat_fill
                        cell_monto.font = cat_font
                        cell_monto.fill = cat_fill
                    elif nivel == "Subcategoria":
                        cell_nombre.font = subcat_font
                        cell_nombre.fill = subcat_fill
                        cell_monto.font = subcat_font
                        cell_monto.fill = subcat_fill
                    elif nivel == "Total":
                        cell_nombre.font = total_font
                        cell_nombre.fill = total_fill
                        cell_monto.font = total_font
                        cell_monto.fill = total_fill
                        cell_nombre.alignment = center_align
                        cell_monto.alignment = right_align

            return True, None
        except Exception as e:
            return False, str(e)

    def to_pdf_resumen_por_cuenta(self, filepath=None):
        import pandas as pd

        if self.df.empty:
            return False, "No hay datos para exportar."
        try:
            if not filepath:
                return False, "No se indicó el nombre del archivo PDF."
            from fpdf import FPDF

            pdf = FPDF(orientation='P', unit='mm', format='Letter')
            pdf.set_auto_page_break(auto=True, margin=18)
            pdf.add_page()

            page_width = pdf.w - 2 * pdf.l_margin

            # Encabezado
            pdf.set_font("Arial", "B", 20)
            pdf.cell(page_width, 16, self.title, ln=True, align='C')
            pdf.set_font("Arial", "I", 13)
            pdf.cell(page_width, 8, f"Proyecto: {self.project_name}", ln=True, align='C')
            pdf.cell(page_width, 8, f"Período: {self.date_range}", ln=True, align='C')
            pdf.ln(8)

            # Colores
            color_header = (79, 129, 189)
            color_text_header = (255, 255, 255)
            color_ingreso = (0, 128, 0)
            color_gasto = (180, 0, 0)
            color_balance = (0, 0, 70)

            # Tabla principal
            col_cuenta = int(page_width * 0.34)
            col_ingreso = int(page_width * 0.22)
            col_gasto = int(page_width * 0.22)
            col_balance = int(page_width * 0.22)

            # Header
            pdf.set_font("Arial", "B", 14)
            pdf.set_fill_color(*color_header)
            pdf.set_text_color(*color_text_header)
            pdf.cell(col_cuenta, 10, "Cuenta", border=1, align='C', fill=True)
            pdf.cell(col_ingreso, 10, "Ingresos", border=1, align='C', fill=True)
            pdf.cell(col_gasto, 10, "Gastos", border=1, align='C', fill=True)
            pdf.cell(col_balance, 10, "Balance", border=1, align='C', fill=True)
            pdf.ln()

            # Filas
            pdf.set_font("Arial", "", 12)
            df = self.df.copy()
            total_ingresos = 0.0
            total_gastos = 0.0
            total_balance = 0.0
            for _, row in df.iterrows():
                cuenta = str(row['Cuenta'])
                ingresos = float(row['Ingresos'])
                gastos = float(row['Gastos'])
                balance = float(row['Balance'])
                total_ingresos += ingresos
                total_gastos += gastos
                total_balance += balance

                pdf.set_text_color(0, 0, 0)
                pdf.cell(col_cuenta, 10, cuenta, border=1, align='L')
                pdf.set_text_color(*color_ingreso)
                pdf.cell(col_ingreso, 10, f"{self.currency} {ingresos:,.2f}", border=1, align='R')
                pdf.set_text_color(*color_gasto)
                pdf.cell(col_gasto, 10, f"{self.currency} {gastos:,.2f}", border=1, align='R')
                pdf.set_text_color(*color_balance)
                pdf.cell(col_balance, 10, f"{self.currency} {balance:,.2f}", border=1, align='R')
                pdf.ln()

            pdf.set_text_color(60, 60, 60)
            pdf.set_font("Arial", "I", 10)
            pdf.cell(page_width, 8, "Montos en verde: Ingresos   |   Montos en rojo y signo menos: Gastos   |   Balance en azul", 0, 1, 'L')

            pdf.output(filepath)
            return True, None
        except Exception as e:
            return False, str(e)


    def to_excel_resumen_por_cuenta(self, filepath):
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if self.df.empty:
            return False, "No hay datos para exportar."
        try:
            df_export = self.df.copy()

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                sheet_name = 'Resumen por Cuenta'
                df_export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)
                worksheet = writer.sheets[sheet_name]

                # ---- Estilos ----
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                ingreso_font = Font(bold=False, color="008000")  # Verde
                gasto_font = Font(bold=False, color="B40000")    # Rojo
                balance_font = Font(bold=False, color="000070")  # Azul
                title_font = Font(bold=True, size=16)
                subtitle_font = Font(italic=True, size=12)
                center_align = Alignment(horizontal='center')
                right_align = Alignment(horizontal='right')
                currency_format = f'"{self.currency}" #,##0.00'
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                # Título y subtítulo
                worksheet.merge_cells('A1:D1')
                title_cell = worksheet['A1']
                title_cell.value = self.title
                title_cell.font = title_font
                title_cell.alignment = center_align

                worksheet.merge_cells('A2:D2')
                project_cell = worksheet['A2']
                project_cell.value = f"Proyecto: {self.project_name}   |   Período: {self.date_range}"
                project_cell.font = subtitle_font
                project_cell.alignment = center_align

                # Encabezados
                header_row = worksheet[5]
                for cell in header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # Ajuste de ancho de columna
                for col_idx, col_name in enumerate(df_export.columns, 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = max(len(str(cell.value)) for cell in worksheet[column_letter] if cell.value)
                    worksheet.column_dimensions[column_letter].width = max(max_length, 15) + 2

                # Formato de filas
                for idx, row in enumerate(df_export.itertuples(), start=6):
                    # Cuenta
                    cell_cuenta = worksheet[f"A{idx}"]
                    cell_cuenta.alignment = Alignment(horizontal='left')
                    cell_cuenta.border = thin_border

                    # Ingresos
                    cell_ingreso = worksheet[f"B{idx}"]
                    cell_ingreso.number_format = currency_format
                    cell_ingreso.font = ingreso_font
                    cell_ingreso.alignment = right_align
                    cell_ingreso.border = thin_border

                    # Gastos
                    cell_gasto = worksheet[f"C{idx}"]
                    cell_gasto.number_format = currency_format
                    cell_gasto.font = gasto_font
                    cell_gasto.alignment = right_align
                    cell_gasto.border = thin_border

                    # Balance
                    cell_balance = worksheet[f"D{idx}"]
                    cell_balance.number_format = currency_format
                    cell_balance.font = balance_font
                    cell_balance.alignment = right_align
                    cell_balance.border = thin_border

            return True, None
        except Exception as e:
            return False, str(e)

    def dashboard_to_pdf(self, filepath, figures, order=None):
        try:
            if not figures or not isinstance(figures, dict):
                return False, "No hay figuras para exportar."
            pdf = PDF(
                orientation='L',
                unit='mm',
                format='A4',
                title=self.title,
                project_name=self.project_name,
                date_range=self.date_range
            )
            default_order = ['ivg', 'gastos_cat', 'ingresos_cat', 'gastos_subcat', 'ingresos_subcat']
            keys = order or [k for k in default_order if k in figures] or list(figures.keys())
            exported = 0
            for key in keys:
                fig = figures.get(key)
                if fig is None:
                    continue
                try:
                    from matplotlib.figure import Figure
                    if not isinstance(fig, Figure) and hasattr(fig, "figure"):
                        fig = fig.figure
                except Exception:
                    pass
                with io.BytesIO() as buf:
                    try:
                        fig.savefig(buf, format='png', dpi=200, bbox_inches='tight')
                    except Exception:
                        continue
                    buf.seek(0)
                    pdf.add_page()
                    img_w = pdf.w - 2 * pdf.l_margin
                    pdf.image(buf, x=pdf.l_margin, y=None, w=img_w)
                    exported += 1
            if exported == 0:
                return False, "No se encontró ninguna Figure válida en 'figures'."
            pdf.output(filepath)
            return True, None
        except Exception as e:
            return False, str(e)